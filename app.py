import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import pytz
import re

st.set_page_config(page_title="2026 여름신앙학교 스케줄", page_icon="📅", layout="centered")

FILE_PATH = '2026 여름신앙학교 데일리 스케줄(최종).xlsx'

# 📢 실시간 공지사항 로드 함수
def load_notice_data():
    try:
        df_notice = pd.read_excel(FILE_PATH, sheet_name='공지사항')
        if df_notice is not None and not df_notice.empty:
            title = str(df_notice.iloc[0, 0]) if pd.notna(df_notice.iloc[0, 0]) else ""
            content = str(df_notice.iloc[0, 1]) if len(df_notice.columns) > 1 and pd.notna(df_notice.iloc[0, 1]) else ""
            return title, content
    except Exception:
        pass
    return None, None

def load_excel_unmerged(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        
        merged_info = []
        for rng in list(ws.merged_cells.ranges):
            top_left_value = ws.cell(row=rng.min_row, column=rng.min_col).value
            merged_info.append((rng, top_left_value))
            
        for rng, _ in merged_info:
            ws.unmerge_cells(str(rng))
            
        for rng, val in merged_info:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    ws.cell(row=r, column=c).value = val
                    
        data = list(ws.values)
        if not data:
            return None
        return pd.DataFrame(data)
    except Exception:
        return None

def load_contacts_data():
    teacher_df = None
    student_df = None
    
    try:
        teacher_df = pd.read_excel(FILE_PATH, sheet_name='교사 비상연락망')
        if teacher_df is not None and not teacher_df.empty:
            teacher_df.columns = teacher_df.columns.str.strip()
    except Exception:
        pass
        
    try:
        student_df = pd.read_excel(FILE_PATH, sheet_name='학생 비상연락망')
        if student_df is not None and not student_df.empty:
            student_df.columns = student_df.columns.str.strip()
            sort_cols = []
            if '학년' in student_df.columns: sort_cols.append('학년')
            if '이름' in student_df.columns: sort_cols.append('이름')
            if sort_cols:
                student_df = student_df.sort_values(by=sort_cols, ascending=True).reset_index(drop=True)
    except Exception:
        pass
        
    return teacher_df, student_df

# 🛠️ 시간 텍스트를 분 단위(0~1439분)로 변환하는 정밀 함수
def parse_time_to_minutes(time_str):
    time_str = str(time_str).strip()
    match = re.search(r'(\d{1,2})\s*:\s*(\d{2})', time_str)
    if match:
        h, m = int(match.group(1)), int(match.group(2))
        return h * 60 + m
    return None

# 🛠️ [정밀 수정] 한국 시각(KST) 및 분 단위 시각 정밀 매칭
def get_current_day_and_time(available_days, time_data):
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    today_str = now.strftime("%Y-%m-%d")
    current_minutes = now.hour * 60 + now.minute
    
    target_day = "DAY1"
    if "08-01" in today_str:
        target_day = "DAY2"
    elif "08-02" in today_str:
        target_day = "DAY3"
        
    if target_day not in available_days and len(available_days) > 0:
        target_day = available_days[0]
        
    selected_time = None
    if target_day in time_data:
        day_items = time_data[target_day]
        
        for item in day_items:
            t_raw = item["time"]
            t_min = parse_time_to_minutes(t_raw)
            
            if t_min is not None and t_min <= current_minutes:
                selected_time = t_raw
            elif t_min is not None and t_min > current_minutes:
                break
                
        if not selected_time and len(day_items) > 0:
            selected_time = day_items[0]["time"]
            
    return target_day, selected_time, now.strftime("%Y년 %m월 %d일 %H시 %M분")

# 메인 제목 및 공지 배너
st.title("📅 여름신앙학교 종합 안내")

notice_title, notice_content = load_notice_data()
if notice_title and notice_title.lower() != 'nan':
    st.error(f"### {notice_title}\n{notice_content if notice_content and notice_content.lower() != 'nan' else ''}")

try:
    df_raw = load_excel_unmerged(FILE_PATH, '타임테이블')
    
    if df_raw is None:
        st.error("엑셀 파일에서 '타임테이블' 시트를 찾을 수 없습니다. 파일명을 확인해 주세요.")
    else:
        header_row_idx = 2
        for idx, row in df_raw.iterrows():
            if str(row.iloc[0]).strip() == 'TIME':
                header_row_idx = idx
                break
                
        header_row = df_raw.iloc[header_row_idx]
        people = [str(val).strip() for val in header_row.iloc[1:].dropna().values if str(val).strip()]
        df_body = df_raw.iloc[header_row_idx + 1:].copy()
        
        main_tab1, main_tab2 = st.tabs(["📅 신앙학교 스케줄", "📞 비상연락망"])
        
        # =========================================================
        # TAB 1: 신앙학교 스케줄
        # =========================================================
        with main_tab1:
            st.subheader("🗓️ 일정표 확인")
            
            schedule_options = [
                "🕒 현재 시간대 스케줄 (실시간)", 
                "🔍 선생님/교사 이름 검색", 
                "🌟 전체 스케줄 보기", 
                "🖨️ 인쇄용 스케줄 (A4 최적화)"
            ]
            selected_schedule = st.selectbox("👀 확인할 스케줄 항목을 선택하세요:", schedule_options, key="schedule_select")
            
            # 1-1. 현재 시간대 스케줄 (분 단위 정밀 수식 매칭)
            if selected_schedule == "🕒 현재 시간대 스케줄 (실시간)":
                time_data = {}
                current_day = "DAY1"
                
                for idx in range(len(df_body)):
                    row = df_body.iloc[idx]
                    time_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    if "DAY" in time_val.upper():
                        current_day = time_val
                        continue
                    if time_val.lower() in ["nan", "none", ""]: 
                        continue
                        
                    tasks = {}
                    has_task = False
                    for person in people:
                        p_idx = list(header_row).index(person)
                        task_val = str(row.iloc[p_idx]).strip() if pd.notna(row.iloc[p_idx]) else ""
                        if task_val.lower() in ["nan", "none"]: task_val = ""
                        tasks[person] = task_val
                        if task_val: has_task = True
                        
                    if has_task:
                        if current_day not in time_data:
                            time_data[current_day] = []
                        time_data[current_day].append({"time": time_val, "tasks": tasks})
                        
                days = list(time_data.keys())
                
                if days:
                    auto_day, auto_time, formatted_now = get_current_day_and_time(days, time_data)
                    
                    day_idx = days.index(auto_day) if auto_day in days else 0
                    selected_day = st.selectbox("📅 날짜(DAY) 선택:", days, index=day_idx, key="time_day_select")
                    
                    day_times = [item["time"] for item in time_data[selected_day]]
                    time_idx = day_times.index(auto_time) if auto_time in day_times else 0
                    selected_time = st.selectbox("⏰ 시간 선택:", day_times, index=time_idx, key="time_hour_select")
                    
                    st.divider()
                    st.markdown(f"### 📍 {selected_day} | {selected_time}")
                    st.caption(f"💡 현재 한국 시각 기준 ({formatted_now}) 진행 일정입니다.")
                    
                    for item in time_data[selected_day]:
                        if item["time"] == selected_time:
                            for person, task in item["tasks"].items():
                                if task:
                                    st.write(f"- **{person}**: {task}")
                                else:
                                    st.write(f"- **{person}**: (공란/휴식)")
                            break

            # 1-2. 선생님 이름 검색
            elif selected_schedule == "🔍 선생님/교사 이름 검색":
                st.markdown("### 🔍 교사/봉사자 개인 일정 빠른 검색")
                search_person = st.selectbox("선생님 이름을 선택하세요:", people, key="person_search_dropdown")
                
                if search_person:
                    person_col_idx = list(header_row).index(search_person)
                    schedule_data = []
                    current_day = "DAY1"
                    
                    for idx in range(len(df_body)):
                        row = df_body.iloc[idx]
                        time_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        if "DAY" in time_val.upper():
                            current_day = time_val
                            continue
                        task_val = str(row.iloc[person_col_idx]).strip() if pd.notna(row.iloc[person_col_idx]) else ""
                        if task_val.lower() in ["nan", "none"]: task_val = ""
                        if time_val.lower() in ["nan", "none"]: time_val = ""
                        
                        if time_val or task_val:
                            schedule_data.append({"DAY": current_day, "시간": time_val, "담당 업무": task_val})
                            
                    df_schedule = pd.DataFrame(schedule_data)
                    st.divider()
                    st.markdown(f"### ✅ **{search_person}** 선생님 일정 요약")
                    
                    days = df_schedule["DAY"].unique()
                    if len(days) > 0:
                        day_tabs = st.tabs(list(days))
                        for i, day in enumerate(days):
                            with day_tabs[i]:
                                day_df = df_schedule[df_schedule["DAY"] == day]
                                st.dataframe(day_df[["시간", "담당 업무"]], hide_index=True, use_container_width=True)

            # 1-3. 전체 스케줄 보기
            elif selected_schedule == "🌟 전체 스케줄 보기":
                st.caption("👈 모바일에서는 표를 좌우로 스크롤하여 모든 인물을 확인할 수 있습니다.")
                all_data = []
                current_day = "DAY1"
                
                for idx in range(len(df_body)):
                    row = df_body.iloc[idx]
                    time_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    if "DAY" in time_val.upper():
                        current_day = time_val
                        continue
                    if time_val.lower() in ["nan", "none"]: time_val = ""
                    
                    row_dict = {"DAY": current_day, "시간": time_val}
                    has_task = False
                    for person in people:
                        p_idx = list(header_row).index(person)
                        task_val = str(row.iloc[p_idx]).strip() if pd.notna(row.iloc[p_idx]) else ""
                        if task_val.lower() in ["nan", "none"]: task_val = ""
                        row_dict[person] = task_val
                        if task_val: has_task = True
                        
                    if time_val or has_task:
                        all_data.append(row_dict)
                        
                df_all = pd.DataFrame(all_data)
                days = df_all["DAY"].unique()
                if len(days) > 0:
                    day_tabs = st.tabs(list(days))
                    for i, day in enumerate(days):
                        with day_tabs[i]:
                            day_df = df_all[df_all["DAY"] == day].drop(columns=["DAY"])
                            st.dataframe(day_df, hide_index=True, use_container_width=False)

            # 1-4. 인쇄용 스케줄
            elif selected_schedule == "🖨️ 인쇄용 스케줄 (A4 최적화)":
                st.success("웹사이트 화면 제약 없이 여러 장을 완벽하게 인쇄하려면 아래의 **[다운로드]** 버튼을 눌러 파일을 받아주세요!")
                print_target = st.selectbox("출력할 대상:", ["전체 스케줄"] + people, key="print_target_select")
                
                all_data = []
                current_day = "DAY1"
                
                for idx in range(len(df_body)):
                    row = df_body.iloc[idx]
                    time_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    if "DAY" in time_val.upper():
                        current_day = time_val
                        continue
                    if time_val.lower() in ["nan", "none"]: time_val = ""
                    
                    row_dict = {"DAY": current_day, "시간": time_val}
                    has_task = False
                    
                    for person in people:
                        p_idx = list(header_row).index(person)
                        task_val = str(row.iloc[p_idx]).strip() if pd.notna(row.iloc[p_idx]) else ""
                        if task_val.lower() in ["nan", "none"]: task_val = ""
                        row_dict[person] = task_val
                        if task_val: has_task = True
                        
                    if time_val or has_task:
                        all_data.append(row_dict)
                        
                df_all = pd.DataFrame(all_data)
                
                if print_target != "전체 스케줄":
                    df_print = df_all[["DAY", "시간", print_target]]
                    df_print = df_print[df_print[print_target] != ""] 
                else:
                    df_print = df_all
                    
                days = df_print["DAY"].unique()
                
                html_content = """
                <!DOCTYPE html>
                <html>
                <head>
                <meta charset="utf-8">
                <title>스케줄 인쇄</title>
                <style>
                    body { font-family: 'Malgun Gothic', sans-serif; padding: 20px; }
                    table { width: 100%; border-collapse: collapse; margin-bottom: 40px; font-size: 11pt; }
                    th, td { border: 1px solid #000; padding: 10px; text-align: center; }
                    th { background-color: #e6e6e6; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
                    tr { page-break-inside: avoid; }
                    thead { display: table-header-group; }
                    h2 { page-break-before: always; margin-bottom: 10px; }
                    h2:first-of-type { page-break-before: auto; }
                    @page { size: auto; margin: 15mm; }
                </style>
                </head>
                <body>
                """
                
                for day in days:
                    day_df = df_print[df_print["DAY"] == day].drop(columns=["DAY"])
                    html_table = day_df.to_html(index=False, escape=False)
                    html_content += f"<h2>{day} - {print_target}</h2>\n"
                    html_content += html_table + "\n"
                    
                html_content += "</body></html>"
                
                st.download_button(
                    label="📥 완벽 인쇄용 파일 다운로드 (클릭)",
                    data=html_content,
                    file_name=f"여름신앙학교_스케줄_{print_target}.html",
                    mime="text/html"
                )

        # =========================================================
        # TAB 2: 비상연락망
        # =========================================================
        with main_tab2:
            st.subheader("📞 비상연락망 안내")
            teacher_df, student_df = load_contacts_data()
            
            contact_tab1, contact_tab2 = st.tabs(["👨‍🏫 교사 연락망", "🧑‍🎓 학생 연락망"])
            
            with contact_tab1:
                if teacher_df is not None and not teacher_df.empty:
                    for idx, row in teacher_df.iterrows():
                        role = str(row.get('직함', '') if pd.notna(row.get('직함')) else '')
                        name = str(row.get('이름', '') if pd.notna(row.get('이름')) else '')
                        b_name = str(row.get('세례명', '') if pd.notna(row.get('세례명')) else '')
                        phone = str(row.get('전화번호', '') if pd.notna(row.get('전화번호')) else '')
                        
                        with st.container(border=True):
                            c1, c2 = st.columns([1.5, 1.5])
                            with c1:
                                st.markdown(f"**{role}** | **{name}** ({b_name})")
                            with c2:
                                if phone and phone.lower() != 'nan':
                                    clean_p = phone.replace('-', '').strip()
                                    st.markdown(f"📞 [{phone}](tel:{clean_p})")
                                else:
                                    st.caption("번호 없음")
                else:
                    st.info("교사 비상연락망 데이터가 없습니다.")
                    
            with contact_tab2:
                if student_df is not None and not student_df.empty:
                    search_q = st.text_input("🔍 학생 이름 또는 세례명 검색:", "", key="student_search_card")
                    
                    filtered_s = student_df.copy()
                    if search_q:
                        mask = filtered_s.astype(str).apply(lambda row: row.str.contains(search_q, case=False).any(), axis=1)
                        filtered_s = filtered_s[mask]
                    
                    view_mode = st.radio("📱 보기 모드:", ["📱 카드형 (모바일 추천)", "📊 표 전체보기"], horizontal=True)
                    
                    if view_mode == "📊 표 전체보기":
                        st.dataframe(filtered_s, hide_index=True, use_container_width=True)
                    else:
                        st.caption(f"총 {len(filtered_s)}명 | 학년별 접이식 메뉴를 눌러보세요.")
                        
                        if '학년' in filtered_s.columns:
                            grades = filtered_s['학년'].unique()
                            for grade in grades:
                                grade_data = filtered_s[filtered_s['학년'] == grade]
                                with st.expander(f"🎓 **{grade}** ({len(grade_data)}명)", expanded=True):
                                    for idx, s_row in grade_data.iterrows():
                                        s_name = str(s_row.get('이름', ''))
                                        s_bname = str(s_row.get('세례명', ''))
                                        s_gender = str(s_row.get('성별', ''))
                                        s_phone = str(s_row.get('학생 연락처', ''))
                                        p_phone = str(s_row.get('학부모 연락처', ''))
                                        
                                        with st.container(border=True):
                                            st.markdown(f"👤 **{s_name}** ({s_bname}) · {s_gender}")
                                            
                                            col_s, col_p = st.columns(2)
                                            with col_s:
                                                if s_phone and s_phone.lower() != 'nan':
                                                    clean_s = s_phone.replace('-', '').strip()
                                                    st.markdown(f"📱 **학생:** [{s_phone}](tel:{clean_s})")
                                                else:
                                                    st.caption("📱 학생: 번호 없음")
                                                    
                                            with col_p:
                                                if p_phone and p_phone.lower() != 'nan':
                                                    clean_p = p_phone.replace('-', '').strip()
                                                    st.markdown(f"👨‍👩‍👦 **부모님:** [{p_phone}](tel:{clean_p})")
                                                else:
                                                    st.caption("👨‍👩‍👦 부모님: 번호 없음")
                        else:
                            for idx, s_row in filtered_s.iterrows():
                                s_name = str(s_row.get('이름', ''))
                                s_bname = str(s_row.get('세례명', ''))
                                s_phone = str(s_row.get('학생 연락처', ''))
                                p_phone = str(s_row.get('학부모 연락처', ''))
                                
                                with st.container(border=True):
                                    st.markdown(f"👤 **{s_name}** ({s_bname})")
                                    col_s, col_p = st.columns(2)
                                    with col_s:
                                        if s_phone and s_phone.lower() != 'nan':
                                            st.markdown(f"📱 [{s_phone}](tel:{s_phone.replace('-', '').strip()})")
                                    with col_p:
                                        if p_phone and p_phone.lower() != 'nan':
                                            st.markdown(f"👨‍👩‍👦 [{p_phone}](tel:{p_phone.replace('-', '').strip()})")
                else:
                    st.info("학생 비상연락망 데이터가 없습니다.")

except Exception as e:
    st.error(f"오류가 발생했습니다: {e}")
